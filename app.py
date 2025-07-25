from flask import Flask, render_template, request, redirect, url_for
from collections import defaultdict
from datetime import timedelta
from flask import send_file, session
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
from fpdf import FPDF
import shutil

app = Flask(__name__)
app.secret_key = 'your_secret_key'
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def generate_backlog_priority_timetable(df, start_date, slot_times):
    subject_students = df.groupby('Subject Code')['Roll No.'].apply(set).to_dict()
    subject_student_counts = {subj: len(students) for subj, students in subject_students.items()}

    # Prioritize subjects by number of students descending
    sorted_subjects = sorted(subject_students.keys(), key=lambda x: subject_student_counts[x], reverse=True)

    scheduled_subjects = set()
    used_dates = defaultdict(lambda: defaultdict(set))
    timetable = []

    current_date = start_date
    total_slots = len(slot_times)

    for subject in sorted_subjects:
        students = subject_students[subject]

        for day_offset in range(1000):
            exam_day = current_date + timedelta(days=day_offset)
            for slot_index in range(total_slots):
                if not any(s in used_dates[exam_day][slot_index] for s in students):
                    subject_name = df[df['Subject Code'] == subject]['Subject Name'].iloc[0]
                    semesters = sorted(map(str, df[df['Subject Code'] == subject]['Semester'].unique()))
                    branches = sorted(map(str, df[df['Subject Code'] == subject]['Batch Name'].unique()))
                    student_count = len(students)

                    timetable.append({
                        'Exam Date': exam_day.strftime('%d-%m-%Y'),
                        'Exam Time': slot_times[slot_index],
                        'Subject Code': subject,
                        'Subject Name': subject_name,
                        'Semester': ', '.join(semesters),
                        'Branch': ', '.join(branches),
                        'Student Count': student_count
                    })

                    used_dates[exam_day][slot_index].update(students)
                    scheduled_subjects.add(subject)
                    break
            else:
                continue
            break

    final_df = pd.DataFrame(timetable)
    final_df.index += 1
    final_df.reset_index(inplace=True)
    final_df.rename(columns={'index': 'Sr. No.'}, inplace=True)
    return final_df, len(set(final_df['Exam Date']))

# Helper function to copy and populate attendance template

def populate_attendance_sheet(template_path, output_path, timetable_df, registration_df):
    dates = sorted(timetable_df['Exam Date'].unique(), key=lambda x: datetime.strptime(x, "%d-%m-%Y"))

    wb_master = load_workbook(template_path)
    template_ws = wb_master.active

    wb_output = load_workbook(template_path)  # Use base template
    wb_output.remove(wb_output.active)  # Remove default sheet

    for date in dates:
        new_ws = wb_output.create_sheet(title=date)
        current_row = 1
        date_data = timetable_df[timetable_df['Exam Date'] == date]

        for _, row in date_data.iterrows():
            subj_code = row['Subject Code']
            subj_name = row['Subject Name']
            exam_time = row['Exam Time']
            semester = row['Semester']
            branch = row['Branch']

            students = registration_df[registration_df['Subject Code'] == subj_code].copy()
            students = students.sort_values(by='Roll No.')

            for i in range(0, len(students), 36):
                chunk = students.iloc[i:i+36].reset_index(drop=True)
                rows_required = len(chunk) + 3  # 3 blank rows

                # Copy template block
                for r in range(1, 41):  # Assume 40 rows block
                    for c in range(1, 10):
                        src_cell = template_ws.cell(row=r, column=c)
                        dest_cell = new_ws.cell(row=current_row + r - 1, column=c, value=src_cell.value)
                        if src_cell.has_style:
                            dest_cell._style = src_cell._style

                # Fill headers (keep your original red box positions)
                new_ws.cell(row=current_row + 3, column=2).value = subj_code
                new_ws.cell(row=current_row + 3, column=5).value = subj_name
                new_ws.cell(row=current_row + 4, column=2).value = date
                new_ws.cell(row=current_row + 4, column=5).value = exam_time
                new_ws.cell(row=current_row + 5, column=2).value = semester
                new_ws.cell(row=current_row + 5, column=5).value = branch

                # Fill student data
                for idx, student in chunk.iterrows():
                    base = current_row + 8 + idx
                    new_ws.cell(row=base, column=1).value = idx + 1
                    new_ws.cell(row=base, column=2).value = student['Roll No.']
                    new_ws.cell(row=base, column=3).value = student['Student Name']

                # Add blank lines if < 36
                for b in range(3):
                    base = current_row + 8 + len(chunk) + b
                    new_ws.cell(row=base, column=1).value = len(chunk) + b + 1

                # Page break after each block
                new_ws.row_breaks.append(Break(id=current_row + 44))
                current_row += 48  # Padding for next block

    wb_output.save(output_path)


def generate_attendance():
    timetable_path = session.get('generated_file')
    registration_path = session.get('uploaded_file')
    template_path = os.path.join('uploads', 'Attendance Sheet Format - (1).xlsx')
    output_excel = os.path.join('uploads', 'attendance_sheets.xlsx')

    timetable_df = pd.read_excel(timetable_path)
    registration_df = pd.read_excel(registration_path, sheet_name='Master File')

    populate_attendance_sheet(template_path, output_excel, timetable_df, registration_df)

    session['attendance_excel'] = output_excel
    return send_file(output_excel, as_attachment=True)

# Flask route to generate attendance
def generate_attendance():
    timetable_path = session.get('generated_file')
    registration_path = session.get('uploaded_file')
    template_path = os.path.join('uploads', 'Attendance Sheet Format.xlsx')
    output_excel = os.path.join('uploads', 'attendance_sheets.xlsx')

    timetable_df = pd.read_excel(timetable_path)
    registration_df = pd.read_excel(registration_path, sheet_name='Master File')

    populate_attendance_sheet(template_path, output_excel, timetable_df, registration_df)

    session['attendance_excel'] = output_excel
    return send_file(output_excel, as_attachment=True)

@app.route('/', methods=['GET', 'POST'])
def home():
    if request.method == 'POST':
        session['exam_name'] = request.form['exam_name']
        session['start_date'] = request.form['start_date']
        session['slot_count'] = int(request.form['slot_count'])
        slot_times = [request.form.get(f'slot_time_{i+1}', '') for i in range(session['slot_count'])]
        session['slot_times'] = slot_times

        file = request.files.get('file')
        if file:
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(filepath)
            session['uploaded_file'] = filepath
        else:
            return "Please upload a valid Excel file.", 400

        return redirect(url_for('generate'))

    return render_template('home.html')

@app.route('/generate')
def generate():
    filepath = session.get('uploaded_file')
    excel_data = pd.ExcelFile(filepath)
    df = excel_data.parse('Master File')
    start_date = datetime.strptime(session.get('start_date'), '%Y-%m-%d')
    slot_times = session.get('slot_times')
    timetable_df, total_days = generate_backlog_priority_timetable(df, start_date, slot_times)

    # Sort by Exam Date and Time
    timetable_df['Exam Date (Sortable)'] = pd.to_datetime(timetable_df['Exam Date'], format='%d-%m-%Y')
    timetable_df = timetable_df.sort_values(by=['Exam Date (Sortable)', 'Exam Time'])
    timetable_df.drop(columns=['Exam Date (Sortable)'], inplace=True)

    output_path = os.path.join(UPLOAD_FOLDER, 'generated_timetable.xlsx')
    timetable_df.to_excel(output_path, index=False)
    session['generated_file'] = output_path

    return render_template('timetable.html', exam_name=session.get('exam_name'), timetable=timetable_df, total_days=total_days)

@app.route('/download')
def download():
    return send_file(session.get('generated_file'), as_attachment=True)

@app.route('/download-template')
def download_template():
    return send_file(os.path.join(UPLOAD_FOLDER, 'registration_template.xlsx'), as_attachment=True)

@app.route('/generate-attendance')
def call_generate_attendance():
    return generate_attendance()


if __name__ == '__main__':
    app.run(debug=True)
